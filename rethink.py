import pandas as pd
import re
import os
import tabula
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from datetime import datetime

def extract_table_from_text(text, title_pattern):
    """Extracts a table from text based on a title pattern."""
    match = re.search(title_pattern + r"\n(.*?)(?=\n\n|\Z)", text, re.DOTALL)
    if not match:
        return None

    table_text = match.group(1)
    lines = table_text.strip().split('\n')

    if not lines:
        return None

    header = [x.strip() for x in lines[0].split(',')]
    data = []
    for line in lines[1:]:
        row = [x.strip() for x in re.split(r',(?!\s)', line)]
        data.append(row)

    return pd.DataFrame(data, columns=header)

def extract_tables_from_pdf(pdf_path):
    """Extracts tables from a PDF using both tabula and text parsing."""
    try:
        tables = tabula.read_pdf(pdf_path, pages='all', multiple_tables=True)
        if tables:
            return tables
    except Exception as e:
        print(f"Tabula extraction failed: {e}")
        pass

    try:
        with open(pdf_path, "r", encoding="utf-8") as f:
            text = f.read()
    except UnicodeDecodeError:
        with open(pdf_path, "r", encoding="latin-1") as f:
            text = f.read()

    table1 = extract_table_from_text(text, r"A\. II\. DETAILS OF OMO PURCHASE ISSUE")
    table2 = extract_table_from_text(text, r"B\. II\. DETAILS OF OMO SALE ISSUE")

    found_tables = []
    if table1 is not None:
        found_tables.append(table1)
    if table2 is not None:
        found_tables.append(table2)
    return found_tables

# --- Main part of the script ---
path = os.getcwd()
pdf_folder = os.path.join(path, 'data')  # Folder where PDFs are stored
os.makedirs(pdf_folder, exist_ok=True)
pdf_files = sorted([f for f in os.listdir(pdf_folder) if f.endswith(".pdf")])

timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
excel_filename = f'output_{timestamp}.xlsx'

all_dfs = {}  # Dictionary to store all DataFrames from all PDFs

# Process each PDF file in the folder
for pdf_file in pdf_files:
    pdf_file_path = os.path.join(pdf_folder, pdf_file)
    print(f"Processing: {pdf_file_path}")

    extracted_tables = extract_tables_from_pdf(pdf_file_path)

    if extracted_tables:
        for i, df in enumerate(extracted_tables):
            if df.shape[1] > 1 and df.shape[0] > 1:
                values = df.values.flatten()
                print(f"Table {i+1} values: {df.shape[1]}")

                columnheader = df.columns
                print(df.columns)  # Print the column names

                # Check if the values array contains NaN values using pd.isna
                if pd.isna(values).all():
                    continue

                # Remove columns with headers that contain "Unnamed"
                df = df.loc[:, ~df.columns.str.contains('^Unnamed')]

                # Remove rows that contain the value "पे्रस प्रकाशनी PRESS RELEASE"
                df = df[~df.apply(lambda row: row.astype(str).str.contains('पे्रस प्रकाशनी PRESS RELEASE').any(), axis=1)]

                table_name = f"{os.path.basename(pdf_file_path)[:-4]}_Table{i+1}"  # create unique table name
                all_dfs[table_name] = df  # Store in the combined dictionary
    else:
        print(f"No tables found in {pdf_file_path}")

# Save all DataFrames to a single Excel file (each DataFrame in a separate sheet)
if all_dfs:
    with pd.ExcelWriter(excel_filename, engine="openpyxl") as writer:
        for sheet_name, df in all_dfs.items():
            if df is not None and not df.empty:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
    print(f"All tables saved to {excel_filename}")
else:
    print("No tables found in any PDF in the folder.")

# Load the workbook to adjust column widths and alignments
wb = load_workbook(excel_filename)

for sheet in wb.sheetnames:
    ws = wb[sheet]
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
        for cell in col:
            cell.alignment = Alignment(horizontal='center', vertical='center')

# Save the adjusted workbook
wb.save(excel_filename)

print(f'Tables have been successfully extracted and formatted in {excel_filename}')